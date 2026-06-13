"""
Export collected JSONL data to an Excel workbook.

Each data type (ticker, orderbook, ohlcv, etc.) becomes a separate sheet.
Nested fields (like slippage_bps, pairwise_spreads) are flattened into columns.

Usage:
    python -m experiments.export_to_excel data/statarb/20260602_190651
    python -m experiments.export_to_excel data/statarb/20260602_190651 --output my_data.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)


def flatten_record(rec: dict, parent_key: str = "", sep: str = ".") -> dict:
    """Flatten nested dicts into dot-separated keys. Skip large lists."""
    items = {}
    for k, v in rec.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(flatten_record(v, new_key, sep))
        elif isinstance(v, list):
            # Small lists (< 20 items of scalars) → join as string
            if len(v) <= 20 and all(not isinstance(i, (dict, list)) for i in v):
                items[new_key] = ", ".join(str(i) for i in v)
            else:
                items[new_key] = f"[{len(v)} items]"
        else:
            items[new_key] = v
    return items


def load_jsonl(path: Path) -> list[dict]:
    """Load a JSONL file, flatten each record."""
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                records.append(flatten_record(rec))
            except json.JSONDecodeError:
                continue
    return records


def records_to_sheet(ws, records: list[dict]):
    """Write flattened records to an openpyxl worksheet."""
    if not records:
        return

    # Collect all unique keys in order of first appearance
    all_keys = []
    seen = set()
    for rec in records:
        for k in rec:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Header row
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row_idx, rec in enumerate(records, 2):
        for col_idx, key in enumerate(all_keys, 1):
            val = rec.get(key)
            # openpyxl can't handle None well in some cases
            if val is None:
                continue
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns (approximate)
    for col_idx, key in enumerate(all_keys, 1):
        max_len = len(key)
        # Sample a few rows for width
        for row_idx in range(2, min(22, len(records) + 2)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def main():
    parser = argparse.ArgumentParser(description="Export JSONL data to Excel")
    parser.add_argument("input_dir", help="Path to data directory (e.g. data/statarb/20260602_190651)")
    parser.add_argument("--output", "-o", help="Output Excel file path (default: <input_dir>.xlsx)")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.is_dir():
        print(f"ERROR: {input_dir} is not a directory")
        sys.exit(1)

    output_path = args.output or str(input_dir.parent / f"{input_dir.name}.xlsx")

    # Discover signals in both layouts:
    #   - daily-partitioned: <input_dir>/<data_type>/<YYYYMMDD>.jsonl
    #   - legacy flat:       <input_dir>/<data_type>.jsonl
    # Each data type becomes one sheet, concatenating all its daily files.
    signals: dict[str, list[Path]] = {}
    for sub in sorted(input_dir.iterdir()):
        if sub.is_dir():
            parts = sorted(sub.glob("*.jsonl"))
            if parts:
                signals.setdefault(sub.name, []).extend(parts)
        elif sub.suffix == ".jsonl":
            signals.setdefault(sub.stem, []).append(sub)

    if not signals:
        print(f"ERROR: No .jsonl files found in {input_dir}")
        sys.exit(1)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_records = 0
    for data_type in sorted(signals):
        sheet_name = data_type[:31]  # Excel sheet name limit
        print(f"  {sheet_name}: ", end="", flush=True)

        records = []
        for jsonl_path in signals[data_type]:
            records.extend(load_jsonl(jsonl_path))
        if not records:
            print("0 records (skipped)")
            continue

        ws = wb.create_sheet(title=sheet_name)
        records_to_sheet(ws, records)
        total_records += len(records)
        print(f"{len(records)} records")

    wb.save(output_path)
    print(f"\nExported {total_records} records across {len(wb.sheetnames)} sheets")
    print(f"  => {output_path}")


if __name__ == "__main__":
    main()
